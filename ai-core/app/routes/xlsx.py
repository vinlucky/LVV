import json
import logging
from pathlib import Path
from fastapi import APIRouter, HTTPException
from fastapi.responses import StreamingResponse
from pydantic import BaseModel

from ..config import FILE_DIR
from ..database import add_message, get_files_by_conv, update_conversation_title, create_task, update_task
from ..services.actor_critic import run_actor_critic_stream
from ..services.file_service import save_generated_file, _get_conv_dir
from ..services.mode_detector import detect_mode_suggestion

logger = logging.getLogger(__name__)
router = APIRouter(prefix="/xlsx", tags=["xlsx"])

XLSX_SYSTEM_PROMPT = """你是一个专业的 Excel 表格生成 AI。请根据用户提供的主题和数据，生成结构化的 Excel 表格数据。

输出格式（JSON）：
{
  "title": "表格标题",
  "sheets": [
    {
      "name": "工作表名称",
      "headers": ["列1", "列2", "列3"],
      "rows": [
        ["数据1", "数据2", "数据3"],
        ["数据4", "数据5", "数据6"]
      ],
      "column_widths": [15, 20, 15],
      "header_style": {
        "bg_color": "4472C4",
        "font_color": "FFFFFF",
        "bold": true
      }
    }
  ]
}

要求：
1. 生成合理的表格结构和数据
2. 列宽适合内容
3. 数据格式规范（数字、百分比、日期等）
4. 适合直接使用
"""


class XLSXGenerateRequest(BaseModel):
    topic: str
    content: str = ""
    style: str = "default"
    model_override: str | None = None
    conv_id: str | None = None
    mode: str = "xlsx"
    ref_files: list[dict] | None = None
    skip_thinking: bool = False


class XLSXRenderRequest(BaseModel):
    outline: dict
    style: str = "default"
    mode: str = "xlsx"
    conv_id: str | None = None


@router.post("/generate/stream")
async def xlsx_generate_stream(req: XLSXGenerateRequest):
    async def generate():
        try:
            if req.conv_id:
                task_info = create_task("xlsx", req.topic[:200])

            suggestion = detect_mode_suggestion(req.topic, "xlsx")
            if suggestion:
                yield f"data: {json.dumps(suggestion, ensure_ascii=False)}\n\n"
                yield "data: [DONE]\n\n"
                return

            prompt = f"Excel 表格主题：{req.topic}\n参考内容：{req.content or '无'}"

            file_content = None
            if req.ref_files:
                from ..services.file_service import read_text_file
                parts = []
                for f in req.ref_files:
                    fp = f.get("file_path", "")
                    fn = f.get("filename", "")
                    if fp:
                        try:
                            content = read_text_file(fp)
                            parts.append(f"=== 文件: {fn} ===\n{content}")
                        except Exception as e:
                            logger.warning(f"Failed to read ref file {fn}: {e}")
                if parts:
                    file_content = "\n\n".join(parts)

            async for event in run_actor_critic_stream(
                user_message=prompt,
                system_prompt_override=XLSX_SYSTEM_PROMPT,
                conv_id=req.conv_id,
                mode="xlsx",
                file_content=file_content,
                skip_thinking=req.skip_thinking,
            ):
                if event.get("type") == "actor_done" and event.get("title") and req.conv_id:
                    try:
                        update_conversation_title(req.conv_id, event["title"], mode="xlsx")
                        event["title_generated"] = True
                    except Exception:
                        pass
                if event.get("type") == "complete":
                    output = event.get("output", "")
                    if req.conv_id:
                        add_message(req.conv_id, "actor", output)
                    try:
                        update_task(task_info["task_id"], "completed", output[:500])
                    except Exception:
                        pass

                    outline = None
                    try:
                        import re
                        json_match = re.search(r'\{[\s\S]*"title"[\s\S]*\}', output)
                        if json_match:
                            outline = json.loads(json_match.group())
                    except Exception:
                        pass

                    if outline:
                        event["outline"] = outline
                        yield f"data: {json.dumps({'type': 'preview_ready', 'outline': outline}, ensure_ascii=False)}\n\n"

                        xlsx_path = _generate_xlsx_from_outline(outline, req.conv_id)
                        if xlsx_path:
                            xlsx_file = {
                                "filename": xlsx_path.name,
                                "file_name": xlsx_path.stem,
                                "file_path": str(xlsx_path),
                                "relative_path": str(xlsx_path.relative_to(FILE_DIR.parent)),
                                "file_format": "xlsx",
                                "file_size": xlsx_path.stat().st_size if xlsx_path.exists() else 0,
                                "conv_id": req.conv_id,
                                "download_url": f"/files/download/{str(xlsx_path.relative_to(FILE_DIR.parent)).replace(chr(92), '/')}",
                            }
                            from ..database import save_file_record
                            save_file_record(
                                filename=xlsx_path.name,
                                stored_filename=xlsx_path.name,
                                file_path=str(xlsx_path),
                                file_format="xlsx",
                                file_size=xlsx_path.stat().st_size if xlsx_path.exists() else 0,
                                conv_id=req.conv_id,
                                mode="xlsx",
                                is_generated=True,
                            )
                            event["file"] = xlsx_file
                            yield f"data: {json.dumps({'type': 'file_generated', 'file': xlsx_file}, ensure_ascii=False)}\n\n"

                            if req.conv_id:
                                conv_files = get_files_by_conv(req.conv_id)
                                add_message(req.conv_id, "generated_files", json.dumps(conv_files, ensure_ascii=False))

                    outline_text = json.dumps(outline, ensure_ascii=False) if outline else output
                    outline_md = save_generated_file(
                        content=outline_text,
                        filename=f"Excel大纲_{req.topic[:20]}",
                        file_format="md",
                        conv_id=req.conv_id,
                        mode="xlsx",
                    )
                    event["outline_file"] = outline_md

                yield f"data: {json.dumps(event, ensure_ascii=False)}\n\n"
            yield "data: [DONE]\n\n"
        except Exception as e:
            logger.error(f"XLSX stream error: {e}")
            yield f"data: {json.dumps({'type': 'error', 'message': str(e)}, ensure_ascii=False)}\n\n"
            yield "data: [DONE]\n\n"

    return StreamingResponse(generate(), media_type="text/event-stream",
                             headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"})


def _generate_xlsx_from_outline(outline: dict, conv_id: str | None = None) -> Path | None:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        wb = Workbook()
        wb.remove(wb.active)

        sheets = outline.get("sheets", [])
        if not sheets:
            sheets = [{"name": "Sheet1", "headers": outline.get("headers", []), "rows": outline.get("rows", [])}]

        for sheet_data in sheets:
            ws = wb.create_sheet(title=sheet_data.get("name", "Sheet1")[:31])

            header_style = sheet_data.get("header_style", {})
            bg_color = header_style.get("bg_color", "4472C4")
            font_color = header_style.get("font_color", "FFFFFF")
            bold = header_style.get("bold", True)

            header_fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type="solid")
            header_font = Font(bold=bold, color=font_color, size=11)
            header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            thin_border = Border(
                left=Side(style="thin", color="D9D9D9"),
                right=Side(style="thin", color="D9D9D9"),
                top=Side(style="thin", color="D9D9D9"),
                bottom=Side(style="thin", color="D9D9D9"),
            )

            headers = sheet_data.get("headers", [])
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment
                cell.border = thin_border

            rows = sheet_data.get("rows", [])
            data_font = Font(size=11)
            data_alignment = Alignment(vertical="center", wrap_text=True)
            for row_idx, row_data in enumerate(rows, 2):
                for col_idx, value in enumerate(row_data, 1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    cell.font = data_font
                    cell.alignment = data_alignment
                    cell.border = thin_border
                    if row_idx % 2 == 0:
                        cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

            column_widths = sheet_data.get("column_widths", [])
            for col_idx, width in enumerate(column_widths, 1):
                ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = width
            if not column_widths:
                for col_idx in range(1, len(headers) + 1):
                    max_len = len(str(headers[col_idx - 1])) if col_idx <= len(headers) else 10
                    for row in rows:
                        if col_idx <= len(row) and row[col_idx - 1]:
                            max_len = max(max_len, len(str(row[col_idx - 1])))
                    ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(max_len + 4, 40)

            ws.auto_filter.ref = ws.dimensions
            ws.freeze_panes = "A2"

        conv_dir = _get_conv_dir(conv_id, "xlsx") if conv_id else FILE_DIR / "xlsx"
        conv_dir.mkdir(parents=True, exist_ok=True)

        title = outline.get("title", "spreadsheet")[:50]
        safe_title = "".join(c if c.isalnum() or c in "._- " else "_" for c in title)
        output_path = conv_dir / f"{safe_title}.xlsx"
        wb.save(str(output_path))
        return output_path
    except Exception as e:
        logger.error(f"Failed to generate XLSX: {e}")
        return None


@router.post("/render")
async def xlsx_render(req: XLSXRenderRequest):
    xlsx_path = _generate_xlsx_from_outline(req.outline, req.conv_id)
    if not xlsx_path:
        raise HTTPException(500, "Failed to generate XLSX file")

    return {
        "filename": xlsx_path.name,
        "file_name": xlsx_path.stem,
        "file_path": str(xlsx_path),
        "relative_path": str(xlsx_path.relative_to(FILE_DIR.parent)),
        "file_format": "xlsx",
        "file_size": xlsx_path.stat().st_size,
        "download_url": f"/files/download/{str(xlsx_path.relative_to(FILE_DIR.parent)).replace(chr(92), '/')}",
    }
